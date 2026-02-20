"""
Exportador a Excel según plantilla de seguimiento.
Escribe en el documento de plantilla, en la hoja correcta.
"""
import re
from pathlib import Path
from datetime import datetime

from schema import BoletaTranscrita
from config import (
    PLANTILLA_EXCEL,
    OUTPUT_DIR,
    HOJA_MAQUINARIAS,
    COLUMNAS_MAQUINARIAS,
    BOLETA_TIPO_HOJA,
    TIPO_DESCONOCIDO,
)


def _parsear_horarios(horarios: str | None) -> tuple[str, str, str]:
    """
    Extrae hora inicial, hora final y horas trabajadas del campo Horarios.
    Ej: "8:00, 13:00, 14:00, 18:00 hs." → (08:00, 18:00, 8)
    Ej: "800 1300, 1400-1800 hs." → (08:00, 18:00, 8)
    """
    h_inicial, h_final, horas = "", "", ""
    if not horarios or not str(horarios).strip():
        return h_inicial, h_final, horas

    texto = str(horarios).replace("hs.", "").replace("hs", "").strip()
    numeros = re.findall(r"\d{1,2}:?\d{0,2}", texto)
    numeros_limpios = []
    for n in numeros:
        if ":" in n:
            partes = n.split(":")
            h = int(partes[0]) if partes[0] else 0
            m = int(partes[1]) if len(partes) > 1 and partes[1] else 0
            numeros_limpios.append((h, m))
        else:
            s = n.zfill(4)
            h, m = int(s[:2]), int(s[2:]) if len(s) > 2 else 0
            numeros_limpios.append((h, m))

    if numeros_limpios:
        primera = numeros_limpios[0]
        ultima = numeros_limpios[-1]
        h_inicial = f"{primera[0]:02d}:{primera[1]:02d}"
        h_final = f"{ultima[0]:02d}:{ultima[1]:02d}"
        try:
            hi_min = primera[0] * 60 + primera[1]
            hf_min = ultima[0] * 60 + ultima[1]
            horas_decimal = (hf_min - hi_min) / 60
            horas = str(round(horas_decimal, 1))
        except Exception:
            pass

    return h_inicial, h_final, horas


def _boleta_a_fila_maquinarias(b: BoletaTranscrita) -> list:
    """Convierte una boleta a fila para la hoja SEGUIMIENTO MAQUINARIAS."""
    h_inicial, h_final, horas = _parsear_horarios(b.horarios)
    return [
        b.numero_reporte or "",
        b.fecha_completa() or "",
        b.operador or "",
        b.obra or "",
        b.patente or "",
        b.camion or "",
        h_inicial,
        h_final,
        horas,
        b.observaciones or "",
    ]



def _escribir_bloque_en_hoja(ws, boletas: list[BoletaTranscrita], col_inicio: int = 2) -> None:
    """Escribe una lista de boletas al final de una hoja existente."""
    fila_inicio = 9
    for row in range(9, min(ws.max_row + 2, 6000)):
        if ws.cell(row=row, column=col_inicio).value is not None:
            fila_inicio = row + 1

    for i, b in enumerate(boletas):
        fila = fila_inicio + i
        valores = _boleta_a_fila_maquinarias(b)
        for col_idx, valor in enumerate(valores):
            ws.cell(row=fila, column=col_inicio + col_idx, value=valor)

def exportar_a_excel(
    boletas: list[BoletaTranscrita],
    ruta_salida: Path | None = None,
    hoja: str = HOJA_MAQUINARIAS,
) -> Path:
    """
    Exporta boletas transcritas al documento de plantilla, en la hoja indicada.

    - Usa PLANTILLA_EXCEL como base
    - Escribe en la hoja existente (SEGUIMIENTO MAQUINARIAS, SEGUIMIENTO_ARIDOS, etc.)
    - Las filas se añaden después de los encabezados (fila 8)
    """
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)

    if ruta_salida is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        ruta_salida = OUTPUT_DIR / f"Seguimiento_Maquinarias_{timestamp}.xlsx"

    try:
        import openpyxl

        wb = openpyxl.load_workbook(PLANTILLA_EXCEL, read_only=False)

        if hoja not in wb.sheetnames:
            raise ValueError(
                f"La hoja '{hoja}' no existe en la plantilla. "
                f"Hojas disponibles: {wb.sheetnames}"
            )

        boletas_por_hoja: dict[str, list[BoletaTranscrita]] = {}
        for b in boletas:
            hoja_tipo = BOLETA_TIPO_HOJA.get(b.tipo_boleta or TIPO_DESCONOCIDO, hoja)
            if hoja_tipo not in wb.sheetnames:
                hoja_tipo = hoja
            boletas_por_hoja.setdefault(hoja_tipo, []).append(b)

        for nombre_hoja, lote in boletas_por_hoja.items():
            ws = wb[nombre_hoja]
            _escribir_bloque_en_hoja(ws, lote, col_inicio=2)

        wb.save(ruta_salida)
        return ruta_salida

    except ImportError:
        import pandas as pd

        datos = [_boleta_a_fila_maquinarias(b) for b in boletas]
        df = pd.DataFrame(datos, columns=COLUMNAS_MAQUINARIAS)
        df.to_excel(ruta_salida, sheet_name=hoja[:31], index=False)
        return ruta_salida
    except Exception as e:
        raise RuntimeError(
            f"No se pudo escribir en la plantilla. Asegúrate de que existe en "
            f"{PLANTILLA_EXCEL} y que tiene la hoja '{hoja}'. Error: {e}"
        ) from e
