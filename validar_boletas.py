#!/usr/bin/env python3
"""
Valida qué boletas ya están transcritas en la plantilla y cuáles faltan.
- Lee reportes de la plantilla Excel
- OCR robusto sobre cada imagen para extraer N° reporte
- Renombra imágenes con formato: reporte_NUMERO.ext
- Muestra: transcritas vs faltantes
"""
import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

from config import EXTENSIONES_IMAGEN, OCR_CONFIG
from preprocesador import preprocesar_para_ocr

# Plantilla con datos transcritos (priorizar la que tenga contenido)
PLANTILLA_1 = Path(__file__).resolve().parent.parent / "Plantilla" / "Seguimiento Maquinarias-aridos-excavaciones(1).xlsx"
PLANTILLA = Path(__file__).resolve().parent.parent / "Plantilla" / "Seguimiento Maquinarias-aridos-excavaciones.xlsx"
IMAGENES_DIR = Path(__file__).resolve().parent.parent / "Imagenes"


def _normalizar_reporte(s: str | None) -> str | None:
    """Normaliza número de reporte: solo dígitos."""
    if not s or not str(s).strip():
        return None
    t = re.sub(r"[^\d]", "", str(s).strip())
    return t if t else None


def _variantes_reporte(norm: str) -> set[str]:
    """Genera variantes para matching (0250 = 250, 000374 = 374)."""
    s = {norm}
    sin_ceros = norm.lstrip("0") or "0"
    s.add(sin_ceros)
    return s


def extraer_reportes_plantilla(ruta_excel: Path) -> set[str]:
    """Extrae todos los N° Report de la plantilla (columna B, hoja SEGUIMIENTO MAQUINARIAS)."""
    try:
        import openpyxl
    except ImportError:
        print("Instala openpyxl: pip install openpyxl")
        return set()

    if not ruta_excel.exists():
        return set()

    wb = openpyxl.load_workbook(ruta_excel, data_only=True)
    if "SEGUIMIENTO MAQUINARIAS" not in wb.sheetnames:
        wb.close()
        return set()

    ws = wb["SEGUIMIENTO MAQUINARIAS"]
    reportes = set()
    for row in range(9, (ws.max_row or 0) + 1):
        v = ws.cell(row=row, column=2).value  # Col B = N° Report
        if v is not None and str(v).strip():
            norm = _normalizar_reporte(str(v))
            if norm:
                reportes.update(_variantes_reporte(norm))
    wb.close()
    return reportes


def listar_imagenes(directorio: Path) -> list[Path]:
    """Lista archivos de imagen válidos."""
    imagenes = []
    for ext in EXTENSIONES_IMAGEN:
        imagenes.extend(directorio.glob(f"*{ext}"))
    imagenes = [p for p in imagenes if "WhatsApp_files" not in str(p)]
    return sorted(imagenes)


def _extraer_numero_reporte_ocr(textos: list) -> tuple[str | None, float]:
    """Extrae N° reporte de resultados OCR. Acepta números de 3-7 dígitos."""
    mapa = str.maketrans({
        "O": "0", "o": "0", "I": "1", "l": "1", "|": "1",
        "Z": "2", "z": "2", "S": "5", "s": "5", "B": "8", "G": "6", "Q": "0",
    })
    mejor_num, mejor_conf = None, 0.0

    for bbox, texto, conf in textos:
        # Patrón N° 0250, Nº 0170224, N. 250
        m = re.search(r"N[°º\.\s]*(\d{3,7})", texto, re.I)
        if m:
            num = re.sub(r"[^\d]", "", m.group(1))
            if 3 <= len(num) <= 7 and conf > mejor_conf:
                mejor_num, mejor_conf = num, conf

        # Número suelto de 4-7 dígitos (típico en boletas)
        num_limpio = texto.strip().translate(mapa)
        num_limpio = re.sub(r"[^\d]", "", num_limpio)
        if 4 <= len(num_limpio) <= 7:
            # Priorizar si está en posición alta (encabezado)
            y = bbox[0][1] if hasattr(bbox[0], "__len__") else 0
            bonus = 0.1 if y < 200 else 0  # Encabezado
            if conf + bonus > mejor_conf:
                mejor_num, mejor_conf = num_limpio, conf + bonus

    return (mejor_num, mejor_conf) if mejor_num else (None, 0.0)


def ocr_extraer_reporte(ruta_imagen: Path, reader, rapido: bool = False) -> tuple[str | None, float]:
    """
    OCR robusto: prueba varios modos de preprocesamiento y devuelve el mejor N° reporte.
    rapido: solo usa modo raw (más rápido).
    """
    modos = ["raw"] if rapido else OCR_CONFIG.get("preproc_try_modes", ["raw", "contraste", "hibrido", "binario"])
    mejor_num, mejor_conf = None, 0.0

    for modo in modos:
        try:
            img = preprocesar_para_ocr(ruta_imagen, modo=modo, devolver_color=True)
            kwargs = {
                "detail": OCR_CONFIG.get("detail", 1),
                "paragraph": False,
                "width_ths": OCR_CONFIG.get("width_ths", 0.7),
            }
            results = reader.readtext(img, **kwargs)
            results = [(b, t.strip(), c) for b, t, c in results if t.strip() and c >= (OCR_CONFIG.get("min_confidence") or 0)]
            num, conf = _extraer_numero_reporte_ocr(results)
            if num and conf > mejor_conf:
                mejor_num, mejor_conf = num, conf
        except Exception:
            continue

    return (mejor_num, mejor_conf)


def main():
    parser = argparse.ArgumentParser(
        description="Valida boletas: cuáles ya transcritas, cuáles faltan. Renombra imágenes."
    )
    parser.add_argument(
        "--imagenes",
        type=Path,
        default=IMAGENES_DIR,
        help="Carpeta con imágenes de boletas",
    )
    parser.add_argument(
        "--plantilla",
        type=Path,
        default=None,
        help="Excel plantilla (default: Plantilla/...xlsx)",
    )
    parser.add_argument(
        "--renombrar",
        action="store_true",
        help="Renombrar imágenes a reporte_NUMERO.ext",
    )
    parser.add_argument(
        "--solo-numero",
        action="store_true",
        help="Renombrar solo con el número (NUMERO.ext)",
    )
    parser.add_argument(
        "--solo-listar",
        action="store_true",
        help="Solo listar sin ejecutar OCR",
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="Más salida",
    )
    parser.add_argument(
        "--exportar-faltantes",
        type=Path,
        metavar="ARCHIVO",
        help="Guardar lista de faltantes en archivo .txt",
    )
    parser.add_argument(
        "--rapido",
        action="store_true",
        help="OCR solo modo raw (4x más rápido, menos preciso)",
    )
    args = parser.parse_args()

    # Resolver plantilla
    plantilla = args.plantilla
    if plantilla is None:
        plantilla = PLANTILLA_1 if PLANTILLA_1.exists() else PLANTILLA

    if not plantilla.exists():
        print(f"ERROR: No existe plantilla en {plantilla}")
        return 1

    reportes_transcritos = extraer_reportes_plantilla(plantilla)
    print(f"Plantilla: {plantilla.name}")
    print(f"Reportes ya transcritos en plantilla: {len(reportes_transcritos)}")
    if args.verbose and reportes_transcritos:
        print("  Ejemplos:", sorted(list(reportes_transcritos))[:15])

    imagenes = listar_imagenes(args.imagenes)
    if not imagenes:
        print(f"\nNo hay imágenes en {args.imagenes}")
        print("  Coloca las imágenes en boletas/Imagenes/ o indica --imagenes /ruta/carpeta")
        return 1

    print(f"\nImágenes encontradas: {len(imagenes)}")

    if args.solo_listar:
        for p in imagenes:
            print(f"  - {p.name}")
        return 0

    # Inicializar OCR
    print("\nInicializando OCR (EasyOCR)...")
    import easyocr
    gpu = OCR_CONFIG.get("gpu", False)
    reader = easyocr.Reader(OCR_CONFIG.get("languages", ["es", "en"]), gpu=gpu, verbose=False)

    transcritas = []   # (imagen, reporte) ya en plantilla
    faltantes = []     # (imagen, reporte) no están en plantilla
    sin_reporte = []   # imágenes donde no se pudo extraer reporte

    from tqdm import tqdm
    renombrar_al_volar = args.solo_numero or args.renombrar
    for ruta in tqdm(imagenes, desc="Analizando"):
        num, conf = ocr_extraer_reporte(ruta, reader, rapido=args.rapido)
        if not num:
            sin_reporte.append((ruta.name, "-"))
            continue
        norm = _normalizar_reporte(num)
        if norm in reportes_transcritos:
            transcritas.append((ruta, norm))
        else:
            faltantes.append((ruta, norm))
        # Renombrar inmediatamente si se pidió (para no perder progreso)
        if renombrar_al_volar and norm:
            ext = ruta.suffix
            fmt = "{reporte}{ext}" if args.solo_numero else "reporte_{reporte}{ext}"
            nuevo = ruta.parent / fmt.format(reporte=norm, ext=ext)
            if ruta != nuevo and ruta.exists():
                if nuevo.exists():
                    base = nuevo.stem
                    i = 1
                    while nuevo.exists():
                        nuevo = ruta.parent / f"{base}_{i}{ext}"
                        i += 1
                try:
                    ruta.rename(nuevo)
                    if norm in reportes_transcritos:
                        transcritas[-1] = (nuevo, norm)
                    else:
                        faltantes[-1] = (nuevo, norm)
                except Exception as e:
                    tqdm.write(f"  [Renombrar falló] {ruta.name}: {e}")

    # Resumen
    print("\n" + "=" * 60)
    print("RESUMEN")
    print("=" * 60)
    print(f"  YA TRANSCRITAS (en plantilla): {len(transcritas)}")
    print(f"  FALTAN TRANSCRIBIR:            {len(faltantes)}")
    print(f"  SIN REPORTE (OCR no detectó):  {len(sin_reporte)}")

    if transcritas:
        print("\n--- Ya transcritas ---")
        for ruta, rep in transcritas[:20]:
            print(f"  {ruta.name}  →  N° {rep}")
        if len(transcritas) > 20:
            print(f"  ... y {len(transcritas) - 20} más")

    if faltantes:
        print("\n--- FALTAN transcribir ---")
        for ruta, rep in faltantes:
            print(f"  {ruta.name}  →  N° {rep}")

    if sin_reporte:
        print("\n--- Sin reporte detectado ---")
        for nombre, _ in sin_reporte[:15]:
            print(f"  {nombre}")
        if len(sin_reporte) > 15:
            print(f"  ... y {len(sin_reporte) - 15} más")

    # Exportar faltantes
    if args.exportar_faltantes and faltantes:
        with open(args.exportar_faltantes, "w", encoding="utf-8") as f:
            f.write("Boletas que FALTAN transcribir a la plantilla\n")
            f.write("=" * 50 + "\n\n")
            for ruta, rep in faltantes:
                f.write(f"{ruta.name}\tN° {rep}\n")
        print(f"\nLista de faltantes guardada en: {args.exportar_faltantes}")

    # Renombrar (si no se hizo ya al vuelo)
    if (args.renombrar or args.solo_numero) and not renombrar_al_volar:
        print("\n--- Renombrando imágenes ---")
        renombrados = 0
        formato = "{reporte}{ext}" if args.solo_numero else "reporte_{reporte}{ext}"
        for ruta, reporte in transcritas + faltantes:
            ext = ruta.suffix
            nuevo = ruta.parent / formato.format(reporte=reporte, ext=ext)
            if ruta != nuevo:
                if nuevo.exists():
                    base = nuevo.stem
                    i = 1
                    while nuevo.exists():
                        nuevo = ruta.parent / f"{base}_{i}{ext}"
                        i += 1
                ruta.rename(nuevo)
                renombrados += 1
                if args.verbose:
                    print(f"  {ruta.name}  →  {nuevo.name}")
        print(f"Renombradas {renombrados} imágenes.")
    elif renombrar_al_volar and (transcritas or faltantes):
        print(f"\nRenombradas {len(transcritas) + len(faltantes)} imágenes (al procesar).")
        if sin_reporte:
            print(f"No se renombraron {len(sin_reporte)} imágenes (sin reporte detectado).")

    return 0


if __name__ == "__main__":
    sys.exit(main())
