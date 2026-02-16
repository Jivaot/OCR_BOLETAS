#!/usr/bin/env python3
"""
Debug: procesa N imágenes, guarda reportes en reportes_live.txt.
Solo guarda si el largo coincide con plantilla (4, 6 o 7 dígitos).
Uso: python debug_5_imagenes.py [--limit N] [--solo-img]
"""
import argparse
import re
import sys
from pathlib import Path

sys.path.insert(0, str(Path(__file__).resolve().parent))

# Largos válidos según plantilla (0250=4, 000374=6, 0170224=7)
LARGOS_VALIDOS = {4, 6, 7}

IMAGENES_DIR = Path(__file__).resolve().parent.parent / "Imagenes"
OUTPUT_TXT = Path(__file__).resolve().parent / "output" / "reportes_live.txt"
PLANTILLA = Path(__file__).resolve().parent.parent / "Plantilla" / "Seguimiento Maquinarias-aridos-excavaciones(1).xlsx"


def _cargar_reportes_validos() -> tuple[set[str], dict[str, str]]:
    """(set de variantes para match, mapa variante->canónico para output)"""
    if not PLANTILLA.exists():
        return (set(), {})
    try:
        import openpyxl
        wb = openpyxl.load_workbook(PLANTILLA, data_only=True)
        ws = wb["SEGUIMIENTO MAQUINARIAS"]
        variantes = set()
        canonico = {}
        for r in range(9, (ws.max_row or 0) + 1):
            v = ws.cell(row=r, column=2).value
            if v:
                n = re.sub(r"[^\d]", "", str(v).strip())
                if n:
                    variantes.add(n)
                    v0 = n.lstrip("0") or "0"
                    variantes.add(v0)
                    canonico[n] = n
                    canonico[v0] = n
        wb.close()
        return (variantes, canonico)
    except Exception:
        return (set(), {})


def _parece_fecha(num: str) -> bool:
    """Rechaza 010500, 300126 etc (fechas DDMMYY o similares)."""
    if len(num) != 6:
        return False
    d = int(num[:2]) if num[:2].isdigit() else 0
    m = int(num[2:4]) if num[2:4].isdigit() else 0
    return 1 <= d <= 31 and 1 <= m <= 12


def _extraer_todos_candidatos(textos: list) -> list[tuple[str, float]]:
    """Retorna TODOS los candidatos (num, conf) sin tomar solo el primero."""
    mapa = str.maketrans({
        "O": "0", "o": "0", "I": "1", "l": "1", "|": "1",
        "Z": "2", "z": "2", "S": "5", "s": "5", "B": "8", "G": "6", "Q": "0",
    })
    candidatos = []
    for bbox, texto, conf in textos:
        m = re.search(r"N[°º\.\s\?]+(\d{3,7})", texto, re.I)
        if m:
            num = re.sub(r"[^\d]", "", m.group(1))
            if len(num) in LARGOS_VALIDOS and not _parece_fecha(num):
                candidatos.append((num, conf))
        num_limpio = re.sub(r"[^\d]", "", texto.strip().translate(mapa))
        if len(num_limpio) in LARGOS_VALIDOS and not _parece_fecha(num_limpio):
            candidatos.append((num_limpio, conf * 0.8))
    return candidatos


def _extraer_numero(textos: list) -> tuple[str | None, float]:
    """
    Extrae SOLO el número de reporte que está junto a N°.
    Prioriza: 1) N° 0254 en mismo texto  2) número a la derecha de N°
    Rechaza fechas (010500) y números sueltos de otras zonas.
    """
    mapa = str.maketrans({
        "O": "0", "o": "0", "I": "1", "l": "1", "|": "1",
        "Z": "2", "z": "2", "S": "5", "s": "5", "B": "8", "G": "6", "Q": "0",
    })
    posiciones_n = []
    numeros_sueltos = []

    for bbox, texto, conf in textos:
        # 1) N° 0254, N? 0253 en mismo texto - máxima prioridad
        m = re.search(r"N[°º\.\s\?]+(\d{3,7})", texto, re.I)
        if m:
            num = re.sub(r"[^\d]", "", m.group(1))
            if len(num) in LARGOS_VALIDOS and not _parece_fecha(num):
                return (num, conf)
        if re.search(r"N[°º\.\?]|N\s*[º\?]", texto, re.I):
            x_end = bbox[2][0] if len(bbox) > 2 else bbox[1][0]
            posiciones_n.append((x_end, bbox[0][1]))
        num_limpio = texto.strip().translate(mapa)
        num_limpio = re.sub(r"[^\d]", "", num_limpio)
        if len(num_limpio) in LARGOS_VALIDOS and not _parece_fecha(num_limpio):
            numeros_sueltos.append((num_limpio, conf, bbox[0][0], bbox[0][1]))

    # 2) Número más cerca de N° (a la derecha, misma fila)
    candidatos = []
    for num, conf, x_num, y_num in numeros_sueltos:
        for xn, yn in posiciones_n:
            if abs(y_num - yn) < 80 and x_num > xn - 50:
                dist = (x_num - xn) ** 2 + (y_num - yn) ** 2
                candidatos.append((dist, num, conf))
                break
    if candidatos:
        candidatos.sort(key=lambda c: c[0])
        return (candidatos[0][1], candidatos[0][2])
    # Fallback: si hay REPORT/DIA/MES en la imagen, el reporte suele ser 4 dígitos (0254)
    # en la zona superior; evitar 6 dígitos tipo fecha
    for num, conf, x_num, y_num in numeros_sueltos:
        if len(num) == 4 and num.startswith("0"):  # 0254, 0259
            return (num, conf * 0.9)  # confianza reducida
    return (None, 0.0)


def main():
    from config import EXTENSIONES_IMAGEN, OCR_CONFIG
    from preprocesador import preprocesar_para_ocr

    ap = argparse.ArgumentParser()
    ap.add_argument("--limit", type=int, default=5, help="Máx imágenes a procesar")
    ap.add_argument("--solo-img", action="store_true", help="Solo archivos IMG-* (excluir ya renombrados)")
    ap.add_argument("--ocr-alt", action="store_true", help="Usar ocr_alternativo (EasyOCR)")
    ap.add_argument("--ocr-celery", action="store_true", help="Usar worker Celery ocr.by_path (CELERY_BROKER_URL)")
    ap.add_argument("--renombrar", action="store_true", help="Renombrar imágenes a NÚMERO.jpg")
    args = ap.parse_args()

    imagenes = []
    for ext in EXTENSIONES_IMAGEN:
        imagenes.extend(IMAGENES_DIR.glob(f"*{ext}"))
    imagenes = sorted(imagenes)
    if args.solo_img:
        imagenes = [p for p in imagenes if p.name.startswith("IMG-")]
    imagenes = imagenes[: args.limit]

    if not imagenes:
        print(f"No hay imágenes en {IMAGENES_DIR}")
        return 1

    OUTPUT_TXT.parent.mkdir(parents=True, exist_ok=True)
    OUTPUT_TXT.write_text("", encoding="utf-8")

    reportes_validos, canonico = _cargar_reportes_validos()
    reportes_usados = {}  # num -> contador para sufijo _1, _2 en duplicados
    print(f"Procesando {len(imagenes)} imágenes → {OUTPUT_TXT}", flush=True)
    print(f"Reportes válidos (plantilla): {len(reportes_validos)}. {'Renombrar: ON' if args.renombrar else ''}\n")

    use_celery = args.ocr_celery
    if use_celery:
        from ocr_celery import leer_imagen_celery, extraer_reporte_de_respuesta_celery
    elif args.ocr_alt:
        from ocr_alternativo import leer_imagen as ocr_leer
    else:
        import easyocr
        reader = easyocr.Reader(OCR_CONFIG.get("languages", ["es", "en"]), gpu=OCR_CONFIG.get("gpu", False), verbose=False)

    for i, ruta in enumerate(imagenes, 1):
        print(f"[{i}/{len(imagenes)}] {ruta.name}...", end=" ", flush=True)
        num, conf, results = None, 0.0, []
        try:
            if use_celery:
                results, raw = leer_imagen_celery(ruta)
                if isinstance(raw, dict) and raw.get("error"):
                    print(f"→ Celery: {raw['error'][:60]}")
                    results = []
                num = extraer_reporte_de_respuesta_celery(raw) if isinstance(raw, dict) else None
                conf = 0.9 if num else 0.0
                if not num and results:
                    num, conf = _extraer_numero(results)
            elif args.ocr_alt:
                results = ocr_leer(ruta)
                num, conf = _extraer_numero(results)
            else:
                # Robustez: múltiples pasadas, elegir el resultado más consistente
                import cv2 as _cv2
                todos_cand = []
                img = preprocesar_para_ocr(ruta, modo="raw", devolver_color=True)
                for modo_img in [("raw", img), ("180", _cv2.rotate(img, _cv2.ROTATE_180))]:
                    res = reader.readtext(modo_img[1], detail=1, paragraph=False)
                    res = [(b, t.strip(), c) for b, t, c in res if t.strip()]
                    todos_cand.extend(_extraer_todos_candidatos(res))
                    if not results and res:
                        results = res
                for modo_alt in ("contraste", "hibrido"):
                    try:
                        img2 = preprocesar_para_ocr(ruta, modo=modo_alt, devolver_color=True)
                        for rot in (img2, _cv2.rotate(img2, _cv2.ROTATE_180)):
                            res = reader.readtext(rot, detail=1, paragraph=False)
                            res = [(b, t.strip(), c) for b, t, c in res if t.strip()]
                            todos_cand.extend(_extraer_todos_candidatos(res))
                    except Exception:
                        pass
                # Si el filename ya tiene un número (0250.jpg, 0259.jpg) y OCR lo encontró, priorizar
                num_del_nombre = re.sub(r"[^\d]", "", ruta.stem)
                if len(num_del_nombre) in LARGOS_VALIDOS:
                    for n, c in todos_cand:
                        nnorm = n.lstrip("0") or "0"
                        ndn_norm = num_del_nombre.lstrip("0") or "0"
                        if n == num_del_nombre or nnorm == ndn_norm or n == num_del_nombre or num_del_nombre == n:
                            num = num_del_nombre if len(num_del_nombre) >= len(n) else n
                            conf = 0.95
                            break
                # Elegir: el que más se repite (consenso), o mayor confianza
                if num is None and todos_cand:
                    from collections import Counter
                    cnt = Counter(n for n, _ in todos_cand)
                    mejor = max(todos_cand, key=lambda x: (cnt[x[0]], x[1]))
                    num, conf = mejor[0], mejor[1]
            # Si el archivo ya se llama 0250.jpg o 0259.jpg (usuario lo puso), priorizar
            m_dn = re.match(r"^(\d{4,7})(?:_\d+)?$", ruta.stem)
            num_dn = m_dn.group(1) if m_dn and len(m_dn.group(1)) in LARGOS_VALIDOS else None
            if num_dn:  # Archivo ya nombrado (0250.jpg, 0259.jpg): priorizar
                num = num_dn
                num_ok = True
            else:
                num_ok = num and len(num) in LARGOS_VALIDOS
            if num_ok and reportes_validos and not num_dn:
                num_norm = num.lstrip("0") or "0"
                num_ok = num in reportes_validos or num_norm in reportes_validos
                if num_ok:
                    num = canonico.get(num) or canonico.get(num_norm) or num
                elif len(num) == 4:
                    # Coincidencia 1 dígito (0259→0254)
                    num4 = num.zfill(4)
                    for rv in reportes_validos:
                        rv4 = rv.zfill(4) if len(rv) <= 4 else rv[:4]
                        if len(rv4) == 4 and sum(a != b for a, b in zip(rv4, num4)) <= 1:
                            num = canonico.get(rv, rv)
                            num_ok = True
                            break
            elif num_ok:
                pass  # sin plantilla, aceptar
            if num_ok:
                # Sufijo si ya usamos este reporte (evitar 7x mismo número)
                reportes_usados[num] = reportes_usados.get(num, 0) + 1
                sufijo = f"_{reportes_usados[num]-1}" if reportes_usados[num] > 1 else ""  # _1, _2 para duplicados
                nombre_final = f"{num}{sufijo}{ruta.suffix}"
                linea = f"{ruta.name}\tN° {num}\n"
                with open(OUTPUT_TXT, "a", encoding="utf-8") as f:
                    f.write(linea)
                    f.flush()
                if args.renombrar:
                    dest = ruta.parent / nombre_final
                    if ruta != dest:
                        n = 1
                        while dest.exists() and dest != ruta:
                            dest = ruta.parent / f"{num}{sufijo}_{n}{ruta.suffix}"
                            n += 1
                        ruta.rename(dest)
                        ruta = dest
                print(f"→ N° {num} ✓ guardado" + (f" → {nombre_final}" if args.renombrar else ""))
            else:
                textos = [t[:40] for _, t, _ in results[:25]] if results else ["(sin resultados OCR)"]
                print(f"→ rechazado | OCR: {textos[:6]!r}")
                # Guardar en txt para que puedas revisar en vivo
                with open(OUTPUT_TXT, "a", encoding="utf-8") as f:
                    f.write(f"# {ruta.name} (rechazado) OCR:\n")
                    for t in textos[:15]:
                        f.write(f"  {t}\n")
                    f.flush()
        except Exception as e:
            print(f"→ ERROR: {e}")

    print(f"\nArchivo: {OUTPUT_TXT}")
    if OUTPUT_TXT.exists():
        print(OUTPUT_TXT.read_text(encoding="utf-8"))
    return 0


if __name__ == "__main__":
    sys.exit(main())
